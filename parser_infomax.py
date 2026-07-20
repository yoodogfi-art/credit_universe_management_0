# -*- coding: utf-8 -*-
"""
VBA Sub parser_info() 의 파이썬(openpyxl) 구현

원본 로직 요약
--------------
1. "Sheet1" 시트의 B열(종목명, raw)을 기준으로 "처음 등장하는 행"만 남겨서
   종목명 -> code_fn(AI열), sector(AL열), 원본행 매핑을 만든다.
   (4214 -> "일반"으로 검색해서 만든 시트라 공사 등이 섞여 들어갈 수 있음)
2. 쓰레기명(IsTrashName) / 출력 제외 대상(IsExcludedFromOutput) 패턴에 해당하는
   종목명은 제외하고, 나머지를 오름차순 정렬한다.
3. 정렬된 종목 목록을 순회하며 bond(회사채) / CP(기업어음) / corp(무보증사채 등)
   등급, 일자, 등급 일치 여부(T/F), 등급전망(긍정/안정/부정) 개수를 계산한다.
4. 결과를 JSON 파일로 저장한다.

수정 내역
--------
- (버그) 파일 최상단의 `openpyxl.read_xlsx(...)` 삭제 (그런 메서드 없음, import 시
  AttributeError).
- (버그) 계산 로직이 함수 밖으로 빠져나가 있어 parser_info()가 아무것도 안 하고
  None을 반환했음. 전체 로직을 함수 내부로 재통합.
- (가독성) 입력 파일/시트/출력 경로를 모듈 상단 상수로 노출 (기존엔 함수 시그니처
  기본값에 숨어 있어서 파일 맨 아래 `if __name__ == "__main__": parser_info()`를
  봐야만 무슨 파일을 읽는지 알 수 있었음).
- (가독성) VBA arrInfo(row, N) 컬럼 번호를 COL 딕셔너리로 이름 붙임. 컬럼이 뭘
  의미하는지 숫자만 보고는 알 수 없었던 문제 해결.
- (변경) 결과 저장 방식을 엑셀 out_info 시트 -> "parsed_infomax.json" 로 변경.
"""

import json
import openpyxl


# ------------------------------------------------------------------
# 실행 시 사용하는 기본 파일 경로 (여기만 보면 무엇을 읽고 쓰는지 바로 파악 가능)
# ------------------------------------------------------------------
DEFAULT_INPUT_PATH = "info_4214.xlsx"
DEFAULT_SHEET_NAME = "Sheet1"
DEFAULT_OUTPUT_PATH = "parsed_infomax.json"


# ------------------------------------------------------------------
# 등급전망(outlook) 판정에 쓰는 한글 키워드
# ------------------------------------------------------------------
POS_STR = "긍정적"
NEU_STR = "안정적"
NEG_STR = "부정적"
NEG_STR2 = "부정적검토"


# ------------------------------------------------------------------
# VBA arrInfo(row, N) 컬럼 번호 (1-based, VBA 컬럼번호 그대로) -> 이름 매핑
# A=1, B=2, C=3, ... Z=26, AA=27, ... AP=42
# ------------------------------------------------------------------
COL = {
    "name": 2,              # B: 종목명 (raw)
    "bond_rating": 3,        # C: 회사채 기준등급
    "bond_date": 4,          # D: 회사채 등급일자
    "bond_r1": 5,             # E: 회사채 평가사1 등급
    "bond_outlook1": 6,       # F: 회사채 등급전망1
    "bond_r2": 8,             # H: 회사채 평가사2 등급
    "bond_outlook2": 9,       # I: 회사채 등급전망2
    "bond_r3": 11,            # K: 회사채 평가사3 등급
    "bond_outlook3": 12,      # L: 회사채 등급전망3
    "cp_rating": 14,          # N: CP 기준등급
    "cp_date": 15,            # O: CP 등급일자
    "cp_r1": 16,               # P: CP 평가사1 등급
    "cp_r2": 18,               # R: CP 평가사2 등급
    "cp_r3": 20,               # T: CP 평가사3 등급
    "corp_rating": 22,        # V: 무보증사채 기준등급
    "corp_date": 23,          # W: 무보증사채 등급일자
    "corp_r1": 24,             # X: 무보증사채 평가사1 등급
    "corp_outlook1": 25,       # Y: 무보증사채 등급전망1
    "corp_r2": 27,             # AA: 무보증사채 평가사2 등급
    "corp_outlook2": 28,       # AB: 무보증사채 등급전망2
    "corp_r3": 30,             # AD: 무보증사채 평가사3 등급
    "corp_outlook3": 31,       # AE: 무보증사채 등급전망3
    "code_fn": 35,             # AI: code_fn
    "sector": 38,              # AL: 섹터
}


# ------------------------------------------------------------------
# 국내 신용등급 스케일 (좋은 순 -> 나쁜 순)
# 장기: 회사채/무보증사채(corp) 등급, 단기: CP 등급
# ------------------------------------------------------------------
RATING_ORDER_LONG = [
    "AAA",
    "AA+", "AA", "AA-",
    "A+", "A", "A-",
    "BBB+", "BBB", "BBB-",
    "BB+", "BB", "BB-",
    "B+", "B", "B-",
    "CCC", "CC", "C", "D",
]

RATING_ORDER_SHORT = [
    "A1",
    "A2+", "A2", "A2-",
    "A3+", "A3", "A3-",
    "B+", "B", "B-",
    "C", "D",
]


def worst_rating(ratings: list, rating_order: list) -> str:
    """
    ratings 중 실제 존재하는 값들(빈 문자열 제외) 중에서 rating_order 기준
    가장 나쁜(=순위상 가장 뒤쪽) 등급을 반환. 전부 비어있으면 "".
    rating_order에 없는 미지의 등급 문자열이 섞여 있으면 가장 나쁜 것보다도
    더 나쁜 취급을 해서(안전 방향) 그대로 반환.
    """
    present = [r for r in ratings if r != ""]
    if not present:
        return ""

    def rank(r):
        return rating_order.index(r) if r in rating_order else len(rating_order)

    return max(present, key=rank)


def tf_check(base_rating: str, r1: str, r2: str, r3: str) -> str:
    """
    TFCheck 대응 함수 [신용등급 parser]

    기준등급(base_rating)과 r1/r2/r3 중 값이 있는 것들을 모아서,
    하나 이상 존재하고 전부 같으면 'T', 하나라도 다르면 'F', 전부 비어있으면 ''.
    """
    ratings = [r for r in (base_rating, r1, r2, r3) if r != ""]

    if len(ratings) == 0:
        return ""

    first = ratings[0]
    for r in ratings:
        if r != first:
            return "F"

    return "T"


def count_outlook(outlook: str, counts: dict) -> None:
    """
    CountOutlook 대응 함수.
    outlook 문자열이 긍정적/안정적/부정적(검토) 중 무엇인지에 따라
    counts = {"pos": 0, "neu": 0, "neg": 0} 를 누적 갱신.
    """
    if outlook == POS_STR:
        counts["pos"] += 1
    elif outlook == NEU_STR:
        counts["neu"] += 1
    elif outlook in (NEG_STR, NEG_STR2):
        counts["neg"] += 1


def is_excluded_case(s: str) -> bool:
    """
    코드/섹터 매핑 단계에서 제외할 예외 케이스 판정.

    아래 중 하나라도 해당하면 True:
      - '-', '(', ')' 포함
      - 'MBS' 포함
      - '제'가 '차'보다 앞에 나오는 경우
      - '유동화' 포함
      - 숫자 뒤에 '차'가 오는 경우
      - 숫자 뒤에 '호'가 오는 경우
      - '스팔', '스페' 포함
      - 마지막 글자가 '우' 또는 'C'
    """
    t = s.strip()

    if "-" in t or "(" in t or ")" in t:
        return True

    if "MBS" in t:
        return True

    pos_je = t.find("제")
    pos_cha = t.find("차")
    if pos_je != -1 and pos_cha != -1 and pos_je < pos_cha:
        return True

    if "유동화" in t:
        return True

    for i in range(len(t) - 1):
        if t[i].isdigit() and t[i + 1] == "차":
            return True
        if t[i].isdigit() and t[i + 1] == "호":
            return True

    if "스팔" in t or "스페" in t:
        return True

    if t.endswith("우") or t.endswith("C"):
        return True

    return False


def _cell(row, col_idx: int) -> str:
    """arr_info 한 행(row)에서 col_idx(1-based, VBA 컬럼번호) 값을 문자열로 안전하게 추출."""
    v = row[col_idx]
    return str(v).strip() if v is not None else ""


def parser_info(
    input_path: str = DEFAULT_INPUT_PATH,
    sheet_name: str = DEFAULT_SHEET_NAME,
    json_output_path: str = DEFAULT_OUTPUT_PATH,
) -> str:
    """
    info_4214 원본 엑셀(input_path, sheet_name)을 읽어 종목별 신용등급 정보를
    파싱하고 json_output_path 에 JSON으로 저장한다. 저장된 경로를 반환한다.
    """

    # ------------------------------------------------------------------
    # 0) 엑셀 파일 열기
    # ------------------------------------------------------------------
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws_ref = wb[sheet_name]

    # ------------------------------------------------------------------
    # B열 기준 마지막 데이터 행 찾기
    # ------------------------------------------------------------------
    last_row = ws_ref.max_row
    while (
        last_row > 1
        and ws_ref.cell(row=last_row, column=COL["name"]).value in (None, "")
    ):
        last_row -= 1

    # ------------------------------------------------------------------
    # A2:AP 범위를 메모리로 로드
    # (인덱스를 VBA 1-based 컬럼번호와 맞추기 위해 맨 앞에 None 패딩)
    # ------------------------------------------------------------------
    arr_info = []
    for row in ws_ref.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=42):
        arr_info.append([None] + [c.value for c in row])

    # ------------------------------------------------------------------
    # 1) 종목명 -> 코드 / 섹터 / 행번호 매핑 (첫 등장 행만 유지)
    # ------------------------------------------------------------------
    code_a = {}
    sector_a = {}
    row_map_a = {}

    for i, row in enumerate(arr_info):
        raw = _cell(row, COL["name"])
        if raw != "" and raw not in code_a:
            code_a[raw] = _cell(row, COL["code_fn"])
            sector_a[raw] = _cell(row, COL["sector"])
            row_map_a[raw] = i

    # ------------------------------------------------------------------
    # 2) 예외 케이스 제거 후 오름차순 정렬 (VBA: SimpleSort)
    # ------------------------------------------------------------------
    only_a = [key for key in code_a.keys() if not is_excluded_case(key)]
    only_a.sort()

    # ------------------------------------------------------------------
    # 3) 종목별 계산
    # ------------------------------------------------------------------
    out_rows = []

    for name in only_a:
        info_row = arr_info[row_map_a[name]]

        # "CANC" 는 취소된 등급이므로 빈 값 취급
        bond_base_rating = _cell(info_row, COL["bond_rating"])
        corp_base_rating = _cell(info_row, COL["corp_rating"])

        if bond_base_rating == "CANC":
            bond_base_rating = ""
        if corp_base_rating == "CANC":
            corp_base_rating = ""

        # 둘 다 없으면 해당 종목은 결과에서 제외 (VBA: GoTo NextEntity)
        if bond_base_rating == "" and corp_base_rating == "":
            continue

        row_out = {
            "code_info": code_a[name],
            "code_fn": "",             # VBA 원본에서도 항상 빈 문자열
            "sector": sector_a[name],
            "name": name,
        }

        # --- Bond (회사채) ---
        bond_r1 = _cell(info_row, COL["bond_r1"])
        bond_r2 = _cell(info_row, COL["bond_r2"])
        bond_r3 = _cell(info_row, COL["bond_r3"])

        row_out["bond_date"] = _cell(info_row, COL["bond_date"])

        if bond_base_rating == "":
            row_out["bond_rating"] = ""
            row_out["bond_TF"] = ""
        else:
            row_out["bond_TF"] = tf_check(bond_base_rating, bond_r1, bond_r2, bond_r3)
            # 3사 등급이 불일치(F)하면 기준등급 대신 최악등급을 저장
            if row_out["bond_TF"] == "F":
                row_out["bond_rating"] = worst_rating(
                    [bond_base_rating, bond_r1, bond_r2, bond_r3], RATING_ORDER_LONG
                )
            else:
                row_out["bond_rating"] = bond_base_rating

        # bond 등급전망은 기준등급 유무와 상관없이 항상 집계
        bond_counts = {"pos": 0, "neu": 0, "neg": 0}
        for key in ("bond_outlook1", "bond_outlook2", "bond_outlook3"):
            count_outlook(_cell(info_row, COL[key]), bond_counts)
        row_out["bond_pos"] = bond_counts["pos"]
        row_out["bond_neu"] = bond_counts["neu"]
        row_out["bond_neg"] = bond_counts["neg"]

        # --- CP (기업어음) ---
        cp_base_rating = _cell(info_row, COL["cp_rating"])
        cp_r1 = _cell(info_row, COL["cp_r1"])
        cp_r2 = _cell(info_row, COL["cp_r2"])
        cp_r3 = _cell(info_row, COL["cp_r3"])

        row_out["cp_date"] = _cell(info_row, COL["cp_date"])

        if cp_base_rating == "":
            row_out["cp_rating"] = ""
            row_out["cp_TF"] = ""
        else:
            row_out["cp_TF"] = tf_check(cp_base_rating, cp_r1, cp_r2, cp_r3)
            # 3사 등급이 불일치(F)하면 기준등급 대신 최악등급을 저장
            if row_out["cp_TF"] == "F":
                row_out["cp_rating"] = worst_rating(
                    [cp_base_rating, cp_r1, cp_r2, cp_r3], RATING_ORDER_SHORT
                )
            else:
                row_out["cp_rating"] = cp_base_rating

        # --- Corp (무보증사채 등) ---
        corp_r1 = _cell(info_row, COL["corp_r1"])
        corp_r2 = _cell(info_row, COL["corp_r2"])
        corp_r3 = _cell(info_row, COL["corp_r3"])

        row_out["corp_date"] = _cell(info_row, COL["corp_date"])

        if corp_base_rating == "":
            row_out["corp_rating"] = ""
            row_out["corp_TF"] = ""
        else:
            row_out["corp_TF"] = tf_check(corp_base_rating, corp_r1, corp_r2, corp_r3)
            # 3사 등급이 불일치(F)하면 기준등급 대신 최악등급을 저장
            if row_out["corp_TF"] == "F":
                row_out["corp_rating"] = worst_rating(
                    [corp_base_rating, corp_r1, corp_r2, corp_r3], RATING_ORDER_LONG
                )
            else:
                row_out["corp_rating"] = corp_base_rating

        # corp 등급전망도 항상 집계
        corp_counts = {"pos": 0, "neu": 0, "neg": 0}
        for key in ("corp_outlook1", "corp_outlook2", "corp_outlook3"):
            count_outlook(_cell(info_row, COL[key]), corp_counts)
        row_out["corp_pos"] = corp_counts["pos"]
        row_out["corp_neu"] = corp_counts["neu"]
        row_out["corp_neg"] = corp_counts["neg"]

        out_rows.append(row_out)

    # ------------------------------------------------------------------
    # 4) JSON으로 저장
    # ------------------------------------------------------------------
    with open(json_output_path, "w", encoding="utf-8") as f:
        json.dump(out_rows, f, ensure_ascii=False, indent=2)

    print(f"Done. {len(out_rows)} companies. -> {json_output_path}")
    return json_output_path


if __name__ == "__main__":
    parser_info(DEFAULT_INPUT_PATH, DEFAULT_SHEET_NAME, DEFAULT_OUTPUT_PATH)