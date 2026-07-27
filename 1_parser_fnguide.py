# -*- coding: utf-8 -*-
"""
VBA Sub parser_fn() 의 파이썬(openpyxl) 구현

입력: fnguide.xlsx (Sheet: fn_all)
출력: parsed_fnguide.csv

원본 로직 요약
--------------
1. "fn_all" 시트, 6행을 헤더로 하는 7개 블록(각 블록 시작 컬럼: C,H,M,R,W,AB,AG)에서
   값을 읽는다.
2. AL열 값에 'T' 또는 '1'이 포함되어 있으면(use_march) 각 블록의 3,4번째 컬럼
   (즉 3월 실적 계열)을, 아니면 1,2번째 컬럼을 우선순위로 읽는다
   (앞 컬럼이 비어있으면 다음 컬럼으로 폴백).
3. A열(code)은 앞뒤 1글자씩 잘라내고, B열(name)은 그대로 사용.
   쓰레기명/출력제외 패턴에 해당하는 종목은 제외.
4. 7개 블록 값 중 앞 5개로 조건 플래그(C>=5, D<=50, E<=10, F>=10, G/H<=0.75)를
   계산하고, 5개 플래그가 전부 유효할 때만 score = 5 - sum(flags)를 매겨
   internal_rating(AG2/AG3)을 정한다. 하나라도 비어있으면 무조건 AG3.
   (추가) internal_rating이 AG3일 때 그 사유를 ag3_reason 컬럼에 남긴다:
     - "score": 플래그 5개가 다 있고 score를 계산했지만 그 결과가 AG3인 경우
     - "재무정보 미제공": 플래그 중 하나라도 비어 있어 score 자체를 못 낸 경우
5. 결과를 CSV로 저장.
"""

import csv
import openpyxl


# ------------------------------------------------------------------
# 실행 시 사용하는 기본 파일 경로
# ------------------------------------------------------------------
DEFAULT_INPUT_PATH = "fnguide.xlsx"
DEFAULT_SHEET_NAME = "fn_all"
DEFAULT_OUTPUT_PATH = "parsed_fnguide.csv"


# ------------------------------------------------------------------
# 블록 시작 컬럼 (1-based, VBA 컬럼번호): C=3, H=8, M=13, R=18, W=23, AB=28, AG=33
# 헤더는 각 블록 시작 컬럼의 6행 값
# ------------------------------------------------------------------
BLOCK_STARTS = [3, 8, 13, 18, 23, 28, 33]

CODE_COL = 1     # A: code (raw)
NAME_COL = 2      # B: name
FLAG_COL = 38     # AL: 3월/1분기 판정용 플래그 값


def _blank(v) -> bool:
    """VBA의 IsEmpty/IsNull/CStr(v)="" 판정에 대응."""
    return v is None or str(v).strip() == ""


def _to_str(v) -> str:
    return "" if v is None else str(v).strip()


def _to_num(v):
    """숫자로 변환 가능하면 float, 아니면 None. VBA IsNumeric 대응."""
    if _blank(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def is_excluded_from_output(s: str) -> bool:
    """
    IsExcludedFromOutput 대응 함수.
    아래 중 하나라도 해당하면 True:
      - '-', '(', ')' 포함
      - 'MBS' 포함
      - '우B' 포함
      - 숫자 뒤에 '차'가 오는 경우
    """
    t = s

    if "-" in t or "(" in t or ")" in t:
        return True

    if "MBS" in t:
        return True

    if "우B" in t:
        return True

    for i in range(len(t) - 1):
        if t[i].isdigit() and t[i + 1] == "차":
            return True

    return False


def is_trash_name(s: str) -> bool:
    """
    IsTrashName 대응 함수.
    아래 중 하나라도 해당하면 True:
      - '(', ')' 포함
      - 'MBS' 포함
      - 숫자 뒤에 '호'가 오는 경우
      - '스팔', '스페' 포함
      - 마지막 글자가 '우' 또는 'C'
    """
    t = s.strip()

    if "(" in t or ")" in t:
        return True

    if "MBS" in t:
        return True

    for i in range(len(t) - 1):
        if t[i].isdigit() and t[i + 1] == "호":
            return True

    if "스팔" in t or "스페" in t:
        return True

    if t.endswith("우") or t.endswith("C"):
        return True

    return False


def parser_fn(
    input_path: str = DEFAULT_INPUT_PATH,
    sheet_name: str = DEFAULT_SHEET_NAME,
    csv_output_path: str = DEFAULT_OUTPUT_PATH,
) -> str:
    """
    fnguide.xlsx의 fn_all 시트를 읽어 종목별 재무 스코어링 결과를
    csv_output_path 에 CSV로 저장한다. 저장된 경로를 반환한다.
    """

    # ------------------------------------------------------------------
    # 0) 엑셀 파일 열기
    # ------------------------------------------------------------------
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb[sheet_name]

    # ------------------------------------------------------------------
    # B열 기준 마지막 데이터 행 찾기
    # ------------------------------------------------------------------
    last_row = ws.max_row
    while last_row > 6 and ws.cell(row=last_row, column=NAME_COL).value in (None, ""):
        last_row -= 1

    # ------------------------------------------------------------------
    # 6행 헤더 (블록별 이름)
    # ------------------------------------------------------------------
    hdrs = [_to_str(ws.cell(row=6, column=bs).value) for bs in BLOCK_STARTS]

    # ------------------------------------------------------------------
    # A7:AL(lastRow) 범위를 메모리로 로드
    # (인덱스를 VBA 1-based 컬럼번호와 맞추기 위해 맨 앞에 None 패딩)
    # ------------------------------------------------------------------
    arr_data = []
    for row in ws.iter_rows(min_row=7, max_row=last_row, min_col=1, max_col=38):
        arr_data.append([None] + [c.value for c in row])

    # ------------------------------------------------------------------
    # 출력 CSV 컬럼: code, name, 블록 헤더 7개, 플래그 5개, score, internal_rating
    # ------------------------------------------------------------------
    flag_names = ["C>=5", "D<=50", "E<=10", "F>=10", "G/H<=0.75"]
    fieldnames = ["code", "name"] + hdrs + flag_names + ["score", "internal_rating", "ag3_reason"]

    out_rows = []

    for row in arr_data:
        raw_code = _to_str(row[CODE_COL])
        raw_name = _to_str(row[NAME_COL])

        if raw_code == "" or raw_name == "":
            continue
        if is_trash_name(raw_name) or is_excluded_from_output(raw_name):
            continue

        edited_code = raw_code[1:-1] if len(raw_code) > 2 else raw_code

        al_val = _to_str(row[FLAG_COL])
        use_march = ("T" in al_val) or ("1" in al_val)

        vals = []
        all_blank = True
        for bs in BLOCK_STARTS:
            if use_march:
                v = row[bs + 2]
                if _blank(v):
                    v = row[bs + 3]
            else:
                v = row[bs]
                if _blank(v):
                    v = row[bs + 1]

            if _blank(v):
                vals.append("")
            else:
                vals.append(v)
                all_blank = False

        if all_blank:
            continue

        row_out = {"code": edited_code, "name": raw_name}
        for h, v in zip(hdrs, vals):
            row_out[h] = v

        # --- 조건 플래그 계산 ---
        c_val = _to_num(vals[0])
        d_val = _to_num(vals[1])
        e_val = _to_num(vals[2])
        f_val = _to_num(vals[3])
        g_val = _to_num(vals[4])
        h_val = _to_num(vals[5])

        row_out["C>=5"] = (1 if c_val >= 5 else 0) if c_val is not None else ""
        row_out["D<=50"] = (1 if d_val <= 50 else 0) if d_val is not None else ""
        row_out["E<=10"] = (1 if e_val <= 10 else 0) if e_val is not None else ""
        row_out["F>=10"] = (1 if f_val >= 10 else 0) if f_val is not None else ""

        if g_val is not None and h_val is not None and h_val != 0:
            row_out["G/H<=0.75"] = 1 if (g_val / h_val) <= 0.75 else 0
        else:
            row_out["G/H<=0.75"] = ""

        # score는 5개 플래그가 전부 유효(빈 값 없음)할 때만 계산
        all_flags_valid = all(row_out[k] != "" for k in flag_names)
        if all_flags_valid:
            flag_sum = sum(int(row_out[k]) for k in flag_names)
            score = 5 - flag_sum
            row_out["score"] = score
            row_out["internal_rating"] = "AG2" if score in (0, 1, 2) else "AG3"
            # 플래그는 다 있는데 score(3,4,5)로 인해 AG3인 경우
            row_out["ag3_reason"] = "" if row_out["internal_rating"] == "AG2" else "재무기준 미달"
        else:
            row_out["score"] = ""
            row_out["internal_rating"] = "AG3"
            # 플래그 중 하나라도 비어있어서 애초에 score 산출이 불가능했던 경우
            row_out["ag3_reason"] = "재무정보 미제공"

        out_rows.append(row_out)

    # ------------------------------------------------------------------
    # CSV로 저장 (한글 포함 -> utf-8-sig로 저장해야 엑셀에서 인코딩 안 깨짐)
    # ------------------------------------------------------------------
    with open(csv_output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    print(f"Done. {len(out_rows)} companies from fn_all. -> {csv_output_path}")
    return csv_output_path


if __name__ == "__main__":
    parser_fn(DEFAULT_INPUT_PATH, DEFAULT_SHEET_NAME, DEFAULT_OUTPUT_PATH)